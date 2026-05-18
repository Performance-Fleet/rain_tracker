# =========================
# STREAMLIT APP - CHUVA POR MUNICÍPIO
# =========================

import streamlit as st
import pandas as pd
import requests
import time
import unicodedata
from datetime import date, timedelta
from pathlib import Path
from io import BytesIO
from openpyxl.utils import get_column_letter


# =========================
# CONFIGURAÇÃO DA PÁGINA
# =========================

st.set_page_config(
    page_title="Consulta de Chuva por Município",
    page_icon="🌧️",
    layout="wide"
)


# =========================
# CONFIGURAÇÕES GERAIS
# =========================

IBGE_MUNICIPIOS_URL = "https://servicodados.ibge.gov.br/api/v1/localidades/municipios?orderBy=nome"
OPEN_METEO_GEOCODING_URL = "https://geocoding-api.open-meteo.com/v1/search"
OPEN_METEO_HISTORICAL_URL = "https://archive-api.open-meteo.com/v1/archive"

OUTPUT_DIR = Path("dados_chuva")
OUTPUT_DIR.mkdir(exist_ok=True)

BATCH_SIZE_CHUVA = 40
SLEEP_GEOCODING = 0.2
SLEEP_CHUVA = 1


# =========================
# FUNÇÕES AUXILIARES
# =========================

def normalizar_texto(texto):
    if pd.isna(texto):
        return ""

    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join([c for c in texto if not unicodedata.combining(c)])

    return texto


def request_with_retry(url, params=None, max_retries=4, timeout=90):
    for tentativa in range(1, max_retries + 1):
        try:
            response = requests.get(url, params=params, timeout=timeout)

            if response.status_code == 200:
                return response.json()

            st.warning(f"Tentativa {tentativa} falhou. Status: {response.status_code}")
            st.warning(response.text[:500])

        except Exception as e:
            st.warning(f"Tentativa {tentativa} com erro: {e}")

        time.sleep(tentativa * 3)

    raise RuntimeError(f"Falha após múltiplas tentativas na URL: {url}")


def get_nested(dicionario, caminho, default=None):
    atual = dicionario

    for chave in caminho:
        if isinstance(atual, dict) and chave in atual:
            atual = atual[chave]
        else:
            return default

    return atual


def chunk_dataframe(df, size):
    for start in range(0, len(df), size):
        yield df.iloc[start:start + size].copy()


def dataframe_to_csv_bytes(df):
    return df.to_csv(index=False).encode("utf-8-sig")


def limpar_nome_aba(nome):
    nome = str(nome)

    caracteres_invalidos = ["\\", "/", "*", "?", ":", "[", "]"]

    for c in caracteres_invalidos:
        nome = nome.replace(c, "_")

    nome = nome.strip()

    if not nome:
        nome = "Sheet"

    return nome[:31]


def dataframe_to_xlsx_bytes(df, sheet_name="Tabela"):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        nome_aba = limpar_nome_aba(sheet_name)

        df.to_excel(
            writer,
            index=False,
            sheet_name=nome_aba
        )

        worksheet = writer.sheets[nome_aba]

        for idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)

            tamanho_coluna = max(
                len(str(col)) + 2,
                12
            )

            tamanho_coluna = min(tamanho_coluna, 35)

            worksheet.column_dimensions[col_letter].width = tamanho_coluna

    return output.getvalue()


def varias_tabelas_to_xlsx_bytes(tabelas):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        nomes_usados = set()

        for nome_aba, df in tabelas.items():
            nome_limpo = limpar_nome_aba(nome_aba)

            nome_final = nome_limpo
            contador = 1

            while nome_final in nomes_usados:
                sufixo = f"_{contador}"
                nome_final = f"{nome_limpo[:31 - len(sufixo)]}{sufixo}"
                contador += 1

            nomes_usados.add(nome_final)

            df.to_excel(
                writer,
                index=False,
                sheet_name=nome_final
            )

            worksheet = writer.sheets[nome_final]

            for idx, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)

                tamanho_coluna = max(
                    len(str(col)) + 2,
                    12
                )

                tamanho_coluna = min(tamanho_coluna, 35)

                worksheet.column_dimensions[col_letter].width = tamanho_coluna

    return output.getvalue()


# =========================
# 1. BUSCAR MUNICÍPIOS NO IBGE
# =========================

@st.cache_data(show_spinner=False)
def buscar_municipios_ibge():
    dados = request_with_retry(IBGE_MUNICIPIOS_URL)

    linhas = []

    for item in dados:
        codigo_ibge = item.get("id")
        municipio = item.get("nome")

        uf_sigla = get_nested(item, ["microrregiao", "mesorregiao", "UF", "sigla"])
        uf_nome = get_nested(item, ["microrregiao", "mesorregiao", "UF", "nome"])

        regiao_sigla = get_nested(item, ["microrregiao", "mesorregiao", "UF", "regiao", "sigla"])
        regiao_nome = get_nested(item, ["microrregiao", "mesorregiao", "UF", "regiao", "nome"])

        linhas.append({
            "codigo_ibge": codigo_ibge,
            "municipio": municipio,
            "uf": uf_sigla,
            "estado": uf_nome,
            "regiao_sigla": regiao_sigla,
            "regiao": regiao_nome
        })

    df = pd.DataFrame(linhas)

    df["codigo_ibge"] = df["codigo_ibge"].astype(str)
    df["municipio"] = df["municipio"].astype(str)
    df["uf"] = df["uf"].astype(str)
    df["estado"] = df["estado"].astype(str)
    df["regiao"] = df["regiao"].astype(str)

    return df


# =========================
# 2. GEOCODING AUTOMÁTICO
# =========================

def escolher_melhor_resultado_geocoding(resultados, municipio, uf, estado):
    if not resultados:
        return None

    municipio_norm = normalizar_texto(municipio)
    estado_norm = normalizar_texto(estado)

    candidatos_br = [
        r for r in resultados
        if r.get("country_code") == "BR"
    ]

    if not candidatos_br:
        candidatos_br = resultados

    for r in candidatos_br:
        nome_resultado = normalizar_texto(r.get("name"))
        admin1 = normalizar_texto(r.get("admin1"))

        if nome_resultado == municipio_norm and admin1 == estado_norm:
            return r

    for r in candidatos_br:
        admin1 = normalizar_texto(r.get("admin1"))

        if admin1 == estado_norm:
            return r

    for r in candidatos_br:
        nome_resultado = normalizar_texto(r.get("name"))

        if nome_resultado == municipio_norm:
            return r

    return candidatos_br[0]


@st.cache_data(show_spinner=False)
def geocodificar_municipio_cache(municipio, uf, estado):
    termos_busca = [
        f"{municipio}, {uf}, Brasil",
        f"{municipio}, {estado}, Brasil",
        municipio
    ]

    for termo in termos_busca:
        params = {
            "name": termo,
            "count": 20,
            "language": "pt",
            "format": "json",
            "countryCode": "BR"
        }

        data = request_with_retry(
            OPEN_METEO_GEOCODING_URL,
            params=params,
            max_retries=3,
            timeout=30
        )

        resultados = data.get("results", [])

        melhor = escolher_melhor_resultado_geocoding(
            resultados=resultados,
            municipio=municipio,
            uf=uf,
            estado=estado
        )

        if melhor:
            return {
                "latitude": melhor.get("latitude"),
                "longitude": melhor.get("longitude"),
                "timezone": melhor.get("timezone"),
                "geocoding_name": melhor.get("name"),
                "geocoding_admin1": melhor.get("admin1"),
                "geocoding_country": melhor.get("country"),
                "geocoding_found": True
            }

        time.sleep(SLEEP_GEOCODING)

    return {
        "latitude": None,
        "longitude": None,
        "timezone": None,
        "geocoding_name": None,
        "geocoding_admin1": None,
        "geocoding_country": None,
        "geocoding_found": False
    }


def adicionar_coordenadas(df_municipios):
    resultados = []

    total = len(df_municipios)

    barra = st.progress(0)
    status = st.empty()

    for i, row in df_municipios.reset_index(drop=True).iterrows():
        municipio = row["municipio"]
        uf = row["uf"]
        estado = row["estado"]

        status.write(f"Geocodificando {i + 1}/{total}: {municipio}/{uf}")

        geo = geocodificar_municipio_cache(
            municipio=municipio,
            uf=uf,
            estado=estado
        )

        linha = row.to_dict()
        linha.update(geo)

        resultados.append(linha)

        barra.progress((i + 1) / total)

        time.sleep(SLEEP_GEOCODING)

    status.empty()
    barra.empty()

    df_geo = pd.DataFrame(resultados)

    df_geo["latitude"] = pd.to_numeric(df_geo["latitude"], errors="coerce")
    df_geo["longitude"] = pd.to_numeric(df_geo["longitude"], errors="coerce")
    df_geo["timezone"] = df_geo["timezone"].fillna("America/Sao_Paulo")

    return df_geo


# =========================
# 3. BUSCAR CHUVA NA OPEN-METEO
# =========================

def fetch_chuva_batch(df_batch, start_date, end_date, timezone_name):
    latitudes = ",".join(df_batch["latitude"].astype(str).tolist())
    longitudes = ",".join(df_batch["longitude"].astype(str).tolist())

    params = {
        "latitude": latitudes,
        "longitude": longitudes,
        "start_date": start_date,
        "end_date": end_date,
        "daily": "precipitation_sum,rain_sum,precipitation_hours",
        "timezone": timezone_name,
        "precipitation_unit": "mm"
    }

    data = request_with_retry(
        OPEN_METEO_HISTORICAL_URL,
        params=params,
        max_retries=4,
        timeout=120
    )

    if isinstance(data, dict):
        data = [data]

    resultados = []

    for i, item in enumerate(data):
        if i >= len(df_batch):
            continue

        municipio_info = df_batch.iloc[i]
        daily = item.get("daily", {})

        if not daily:
            continue

        temp = pd.DataFrame(daily)

        if temp.empty:
            continue

        temp.insert(0, "codigo_ibge", municipio_info["codigo_ibge"])
        temp.insert(1, "municipio", municipio_info["municipio"])
        temp.insert(2, "uf", municipio_info["uf"])
        temp.insert(3, "estado", municipio_info["estado"])
        temp.insert(4, "regiao", municipio_info["regiao"])
        temp.insert(5, "latitude", municipio_info["latitude"])
        temp.insert(6, "longitude", municipio_info["longitude"])
        temp.insert(7, "timezone", timezone_name)
        temp.insert(8, "geocoding_name", municipio_info["geocoding_name"])
        temp.insert(9, "geocoding_admin1", municipio_info["geocoding_admin1"])

        resultados.append(temp)

    if resultados:
        return pd.concat(resultados, ignore_index=True)

    return pd.DataFrame()


def buscar_chuva_municipios(df_municipios_geo, start_date, end_date):
    bases = []

    df_ok = df_municipios_geo.dropna(subset=["latitude", "longitude"]).copy()

    total_municipios = len(df_ok)

    if total_municipios == 0:
        raise RuntimeError("Nenhum município com latitude/longitude para buscar chuva.")

    processados = 0

    barra = st.progress(0)
    status = st.empty()

    for timezone_name, df_fuso in df_ok.groupby("timezone"):
        status.write(f"Processando fuso: {timezone_name} | Municípios: {len(df_fuso)}")

        for idx, batch in enumerate(chunk_dataframe(df_fuso, BATCH_SIZE_CHUVA), start=1):
            status.write(
                f"Buscando chuva | Fuso: {timezone_name} | Batch {idx} | {len(batch)} municípios"
            )

            df_chuva_batch = fetch_chuva_batch(
                df_batch=batch,
                start_date=start_date,
                end_date=end_date,
                timezone_name=timezone_name
            )

            if not df_chuva_batch.empty:
                bases.append(df_chuva_batch)

            processados += len(batch)
            barra.progress(processados / total_municipios)

            time.sleep(SLEEP_CHUVA)

    status.empty()
    barra.empty()

    if not bases:
        raise RuntimeError("Nenhum dado de chuva foi retornado pela API.")

    chuva = pd.concat(bases, ignore_index=True)

    chuva = chuva.rename(columns={
        "time": "data",
        "precipitation_sum": "chuva_mm",
        "rain_sum": "chuva_liquida_mm",
        "precipitation_hours": "horas_com_chuva"
    })

    chuva["data"] = pd.to_datetime(chuva["data"]).dt.date
    chuva["chuva_mm"] = pd.to_numeric(chuva["chuva_mm"], errors="coerce")
    chuva["chuva_liquida_mm"] = pd.to_numeric(chuva["chuva_liquida_mm"], errors="coerce")
    chuva["horas_com_chuva"] = pd.to_numeric(chuva["horas_com_chuva"], errors="coerce")

    colunas_finais = [
        "codigo_ibge",
        "municipio",
        "uf",
        "estado",
        "regiao",
        "latitude",
        "longitude",
        "timezone",
        "data",
        "chuva_mm",
        "chuva_liquida_mm",
        "horas_com_chuva",
        "geocoding_name",
        "geocoding_admin1"
    ]

    chuva = chuva[colunas_finais].copy()

    return chuva


# =========================
# 4. TABELA PRINCIPAL
# =========================

def montar_tabela_principal(chuva, granularidade, metrica, data_inicio, data_fim):
    df = chuva.copy()

    df["data"] = pd.to_datetime(df["data"])

    data_inicio_dt = pd.to_datetime(data_inicio)
    data_fim_dt = pd.to_datetime(data_fim)

    df = df[
        (df["data"] >= data_inicio_dt) &
        (df["data"] <= data_fim_dt)
    ].copy()

    if df.empty:
        return pd.DataFrame()

    if granularidade == "Dia":
        df["periodo"] = df["data"].dt.strftime("%d/%m")

        tabela = (
            df
            .groupby(
                [
                    "regiao",
                    "uf",
                    "estado",
                    "municipio",
                    "periodo"
                ],
                as_index=False
            )
            .agg(valor=(metrica, "sum"))
        )

        ordem_periodos = (
            df[["data", "periodo"]]
            .drop_duplicates()
            .sort_values("data")["periodo"]
            .tolist()
        )

    else:
        df["inicio_semana"] = df["data"] - pd.to_timedelta(df["data"].dt.weekday, unit="D")
        df["fim_semana"] = df["inicio_semana"] + pd.Timedelta(days=6)

        df["fim_semana_ajustada"] = df["fim_semana"].where(
            df["fim_semana"] <= data_fim_dt,
            data_fim_dt
        )

        df["periodo"] = (
            df["inicio_semana"].dt.strftime("%d/%m")
            + " a "
            + df["fim_semana_ajustada"].dt.strftime("%d/%m")
        )

        tabela = (
            df
            .groupby(
                [
                    "regiao",
                    "uf",
                    "estado",
                    "municipio",
                    "inicio_semana",
                    "periodo"
                ],
                as_index=False
            )
            .agg(valor=(metrica, "sum"))
        )

        ordem_periodos = (
            tabela[["inicio_semana", "periodo"]]
            .drop_duplicates()
            .sort_values("inicio_semana")["periodo"]
            .tolist()
        )

    tabela_pivot = (
        tabela
        .pivot_table(
            index=[
                "regiao",
                "uf",
                "estado",
                "municipio"
            ],
            columns="periodo",
            values="valor",
            aggfunc="sum",
            fill_value=0
        )
        .reset_index()
    )

    colunas_base = ["regiao", "uf", "estado", "municipio"]

    colunas_periodo = [
        col for col in ordem_periodos
        if col in tabela_pivot.columns
    ]

    tabela_pivot = tabela_pivot[colunas_base + colunas_periodo].copy()

    return tabela_pivot


# =========================
# 5. APP STREAMLIT
# =========================

st.title("🌧️ Consulta de Chuva por Município")
st.caption("Base IBGE + geocoding Open-Meteo + histórico de chuva Open-Meteo")


# =========================
# CARREGAR MUNICÍPIOS
# =========================

with st.spinner("Carregando municípios do IBGE..."):
    municipios = buscar_municipios_ibge()


# =========================
# SIDEBAR
# =========================

st.sidebar.header("Filtros")


# =========================
# PERÍODO AUTOMÁTICO
# =========================
# Regra:
# - D-1 como data final
# - semana começa na segunda
# - busca até 4 semanas contando a semana do D-1

d_1 = date.today() - timedelta(days=1)

inicio_semana_d_1 = d_1 - timedelta(days=d_1.weekday())

data_inicio = inicio_semana_d_1 - timedelta(weeks=3)
data_fim = d_1

st.sidebar.info(
    f"Período automático: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
)

granularidade = st.sidebar.radio(
    "Granularidade da tabela principal",
    options=["Dia", "Semana"],
    horizontal=True
)

metrica_tabela = st.sidebar.selectbox(
    "Métrica da tabela principal",
    options=[
        "chuva_mm",
        "chuva_liquida_mm",
        "horas_com_chuva"
    ],
    index=0
)


# =========================
# FILTROS COMEÇANDO VAZIOS
# Se ficar vazio, considera todos.
# =========================

regioes_disponiveis = sorted(municipios["regiao"].dropna().unique().tolist())

regioes_selecionadas = st.sidebar.multiselect(
    "Regiões",
    options=regioes_disponiveis,
    default=[],
    placeholder="Todas"
)

df_filtrado = municipios.copy()

if regioes_selecionadas:
    df_filtrado = df_filtrado[
        df_filtrado["regiao"].isin(regioes_selecionadas)
    ].copy()


ufs_disponiveis = sorted(df_filtrado["uf"].dropna().unique().tolist())

ufs_selecionadas = st.sidebar.multiselect(
    "UFs",
    options=ufs_disponiveis,
    default=[],
    placeholder="Todas"
)

if ufs_selecionadas:
    df_filtrado = df_filtrado[
        df_filtrado["uf"].isin(ufs_selecionadas)
    ].copy()


municipios_disponiveis = sorted(df_filtrado["municipio"].dropna().unique().tolist())

municipios_selecionados = st.sidebar.multiselect(
    "Municípios específicos",
    options=municipios_disponiveis,
    default=[],
    placeholder="Todos"
)

if municipios_selecionados:
    df_filtrado = df_filtrado[
        df_filtrado["municipio"].isin(municipios_selecionados)
    ].copy()


# Agora não tem modo teste.
# Se nenhum filtro for selecionado, roda todos os municípios.
df_execucao = df_filtrado.copy()


# =========================
# RESUMO DOS FILTROS
# =========================

st.subheader("Municípios selecionados")

col1, col2, col3, col4, col5 = st.columns(5)

col1.metric("Regiões", df_execucao["regiao"].nunique())
col2.metric("UFs", df_execucao["uf"].nunique())
col3.metric("Municípios", df_execucao["codigo_ibge"].nunique())
col4.metric("Granularidade", granularidade)
col5.metric("Métrica", metrica_tabela)

st.caption(
    f"Período considerado: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
)

st.info(
    "Se nenhum filtro for selecionado, o app considera todos os municípios disponíveis na base do IBGE."
)

st.dataframe(
    df_execucao,
    use_container_width=True,
    hide_index=True
)


# =========================
# BOTÃO DE EXECUÇÃO
# =========================

executar = st.button("Buscar dados de chuva", type="primary")


if executar:
    if df_execucao.empty:
        st.error("Nenhum município selecionado.")
        st.stop()

    start_date = data_inicio.isoformat()
    end_date = data_fim.isoformat()

    try:
        # =========================
        # GEOCODIFICAÇÃO
        # =========================

        st.divider()
        st.subheader("1. Geocodificação dos municípios")

        municipios_geo = adicionar_coordenadas(df_execucao)

        st.success("Geocodificação finalizada.")

        st.dataframe(
            municipios_geo,
            use_container_width=True,
            hide_index=True
        )

        nao_encontrados = municipios_geo[
            municipios_geo["geocoding_found"] == False
        ].copy()

        if not nao_encontrados.empty:
            st.warning("Alguns municípios não tiveram coordenadas encontradas.")
            st.dataframe(
                nao_encontrados[["codigo_ibge", "municipio", "uf", "estado"]],
                use_container_width=True,
                hide_index=True
            )

        municipios_geo_path = OUTPUT_DIR / "municipios_geocodificados_auto.csv"
        municipios_geo.to_csv(
            municipios_geo_path,
            index=False,
            encoding="utf-8-sig"
        )

        col_down_geo_1, col_down_geo_2 = st.columns(2)

        with col_down_geo_1:
            st.download_button(
                label="Baixar municípios geocodificados CSV",
                data=dataframe_to_csv_bytes(municipios_geo),
                file_name="municipios_geocodificados_auto.csv",
                mime="text/csv"
            )

        with col_down_geo_2:
            st.download_button(
                label="Baixar municípios geocodificados XLSX",
                data=dataframe_to_xlsx_bytes(
                    municipios_geo,
                    sheet_name="Municipios Geocodificados"
                ),
                file_name="municipios_geocodificados_auto.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # =========================
        # BUSCAR CHUVA
        # =========================

        st.divider()
        st.subheader("2. Busca de chuva")

        chuva = buscar_chuva_municipios(
            df_municipios_geo=municipios_geo,
            start_date=start_date,
            end_date=end_date
        )

        st.success("Base de chuva gerada com sucesso.")

        st.dataframe(
            chuva,
            use_container_width=True,
            hide_index=True
        )

        nome_base = f"chuva_municipios_{start_date}_a_{end_date}"

        csv_path = OUTPUT_DIR / f"{nome_base}.csv"
        parquet_path = OUTPUT_DIR / f"{nome_base}.parquet"

        chuva.to_csv(
            csv_path,
            index=False,
            encoding="utf-8-sig"
        )

        try:
            chuva.to_parquet(parquet_path, index=False)
        except Exception:
            parquet_path = None

        col_down_chuva_1, col_down_chuva_2 = st.columns(2)

        with col_down_chuva_1:
            st.download_button(
                label="Baixar base de chuva CSV",
                data=dataframe_to_csv_bytes(chuva),
                file_name=f"{nome_base}.csv",
                mime="text/csv"
            )

        with col_down_chuva_2:
            st.download_button(
                label="Baixar base de chuva XLSX",
                data=dataframe_to_xlsx_bytes(
                    chuva,
                    sheet_name="Base Chuva"
                ),
                file_name=f"{nome_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # =========================
        # TABELA PRINCIPAL
        # =========================

        st.divider()
        st.subheader("3. Tabela principal")

        tabela_principal = montar_tabela_principal(
            chuva=chuva,
            granularidade=granularidade,
            metrica=metrica_tabela,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        if tabela_principal.empty:
            st.warning("Nenhum dado encontrado para montar a tabela principal.")
        else:
            st.caption(
                f"Granularidade: {granularidade} | "
                f"Métrica: {metrica_tabela} | "
                f"Período: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
            )

            st.dataframe(
                tabela_principal,
                use_container_width=True,
                hide_index=True
            )

            col_down_principal_1, col_down_principal_2 = st.columns(2)

            with col_down_principal_1:
                st.download_button(
                    label="Baixar tabela principal CSV",
                    data=dataframe_to_csv_bytes(tabela_principal),
                    file_name=f"tabela_principal_{granularidade.lower()}_{start_date}_a_{end_date}.csv",
                    mime="text/csv"
                )

            with col_down_principal_2:
                st.download_button(
                    label="Baixar tabela principal XLSX",
                    data=dataframe_to_xlsx_bytes(
                        tabela_principal,
                        sheet_name="Tabela Principal"
                    ),
                    file_name=f"tabela_principal_{granularidade.lower()}_{start_date}_a_{end_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # =========================
        # VALIDAÇÃO POR DIA
        # =========================

        st.divider()
        st.subheader("4. Validação por dia")

        validacao_dia = (
            chuva
            .groupby("data", as_index=False)
            .agg(
                qtd_municipios=("codigo_ibge", "nunique"),
                chuva_media_mm=("chuva_mm", "mean"),
                chuva_maxima_mm=("chuva_mm", "max"),
                municipios_com_chuva=("chuva_mm", lambda x: (x > 0).sum())
            )
        )

        st.dataframe(
            validacao_dia,
            use_container_width=True,
            hide_index=True
        )

        col_down_val_dia_1, col_down_val_dia_2 = st.columns(2)

        with col_down_val_dia_1:
            st.download_button(
                label="Baixar validação por dia CSV",
                data=dataframe_to_csv_bytes(validacao_dia),
                file_name=f"{nome_base}_validacao_dia.csv",
                mime="text/csv"
            )

        with col_down_val_dia_2:
            st.download_button(
                label="Baixar validação por dia XLSX",
                data=dataframe_to_xlsx_bytes(
                    validacao_dia,
                    sheet_name="Validacao Dia"
                ),
                file_name=f"{nome_base}_validacao_dia.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # =========================
        # VALIDAÇÃO POR UF E DIA
        # =========================

        st.subheader("5. Validação por UF e dia")

        validacao_uf = (
            chuva
            .groupby(["uf", "data"], as_index=False)
            .agg(
                qtd_municipios=("codigo_ibge", "nunique"),
                chuva_media_mm=("chuva_mm", "mean"),
                chuva_maxima_mm=("chuva_mm", "max"),
                municipios_com_chuva=("chuva_mm", lambda x: (x > 0).sum())
            )
        )

        st.dataframe(
            validacao_uf,
            use_container_width=True,
            hide_index=True
        )

        col_down_val_uf_1, col_down_val_uf_2 = st.columns(2)

        with col_down_val_uf_1:
            st.download_button(
                label="Baixar validação por UF CSV",
                data=dataframe_to_csv_bytes(validacao_uf),
                file_name=f"{nome_base}_validacao_uf.csv",
                mime="text/csv"
            )

        with col_down_val_uf_2:
            st.download_button(
                label="Baixar validação por UF XLSX",
                data=dataframe_to_xlsx_bytes(
                    validacao_uf,
                    sheet_name="Validacao UF"
                ),
                file_name=f"{nome_base}_validacao_uf.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # =========================
        # AGREGADO POR MUNICÍPIO
        # =========================

        st.subheader("6. Agregado por município no período")

        chuva_periodo_municipio = (
            chuva
            .groupby([
                "codigo_ibge",
                "municipio",
                "uf",
                "estado",
                "regiao",
                "latitude",
                "longitude"
            ], as_index=False)
            .agg(
                chuva_total_periodo_mm=("chuva_mm", "sum"),
                chuva_media_dia_mm=("chuva_mm", "mean"),
                chuva_maxima_dia_mm=("chuva_mm", "max"),
                dias_com_chuva=("chuva_mm", lambda x: (x > 0).sum()),
                horas_com_chuva_total=("horas_com_chuva", "sum")
            )
        )

        periodo_csv_path = OUTPUT_DIR / f"{nome_base}_agregado_municipio.csv"

        chuva_periodo_municipio.to_csv(
            periodo_csv_path,
            index=False,
            encoding="utf-8-sig"
        )

        st.dataframe(
            chuva_periodo_municipio,
            use_container_width=True,
            hide_index=True
        )

        col_down_agg_1, col_down_agg_2 = st.columns(2)

        with col_down_agg_1:
            st.download_button(
                label="Baixar agregado por município CSV",
                data=dataframe_to_csv_bytes(chuva_periodo_municipio),
                file_name=f"{nome_base}_agregado_municipio.csv",
                mime="text/csv"
            )

        with col_down_agg_2:
            st.download_button(
                label="Baixar agregado por município XLSX",
                data=dataframe_to_xlsx_bytes(
                    chuva_periodo_municipio,
                    sheet_name="Agregado Municipio"
                ),
                file_name=f"{nome_base}_agregado_municipio.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # =========================
        # DOWNLOAD ÚNICO COM TODAS AS TABELAS
        # =========================

        st.divider()
        st.subheader("7. Baixar todas as tabelas em um único Excel")

        tabelas_excel = {
            "Municipios Selecionados": df_execucao,
            "Municipios Geocodificados": municipios_geo,
            "Base Chuva": chuva,
            "Tabela Principal": tabela_principal,
            "Validacao Dia": validacao_dia,
            "Validacao UF": validacao_uf,
            "Agregado Municipio": chuva_periodo_municipio
        }

        st.download_button(
            label="Baixar arquivo completo XLSX",
            data=varias_tabelas_to_xlsx_bytes(tabelas_excel),
            file_name=f"relatorio_chuva_completo_{granularidade.lower()}_{start_date}_a_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    except Exception as e:
        st.error("Erro durante a execução.")
        st.exception(e)